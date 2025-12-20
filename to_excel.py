from firebase_bd import *
import pandas as pd
import openpyxl

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
data = leer_registro('capacitaciones')



# Paso 1: Convertir a lista de dicts
rows = list(data.values())

# Paso 2: Crear DataFrame
df = pd.DataFrame(rows)

# Paso 3: Exportar a Excel
archivo_excel = r"C:\Users\alain\Downloads\12F3C710 - copia.xlsx"
df.to_excel(archivo_excel, index=False)

# Paso 4: Aplicar formato con openpyxl
wb = load_workbook(archivo_excel)
ws = wb.active

# Estilos básicos
header_font = Font(bold=True, color="FFFFFF")
header_fill = "538ED5"  # Azul suave
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# Aplicar formato a encabezados
for col_idx, cell in enumerate(ws[1], 1):
    cell.font = header_font
    cell.alignment = center_align
    cell.fill = openpyxl.styles.PatternFill(start_color=header_fill, fill_type="solid")
    cell.border = thin_border
    # Ajustar ancho
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = max(15, len(str(cell.value)) + 2)

# Aplicar borde y alineación al resto
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
    for cell in row:
        cell.border = thin_border
        cell.alignment = center_align

# Guardar archivo con formato
wb.save(archivo_excel)

print("✅ Excel exportado con formato:", archivo_excel)
